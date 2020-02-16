import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates

try:
    import kivy
except ImportError:
    raise ImportError("You don't have Kivy installed!")

from kivy.app import App
from kivy.uix.gridlayout import GridLayout
from kivy.uix.button import Button
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.label import Label
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.textinput import TextInput

filename = "Savings.xlsx"
dates = []
totals = []

extrapolation = 2000


def plotValues(*args):
    try:
        wb = openpyxl.load_workbook(filename)
        print(filename)
    except openpyxl.utils.exceptions.InvalidFileException:
        print("Invalid file name!")

        return

    sheet = wb.get_sheet_by_name('Sheet1')

    for i in range(3, sheet.max_row-1):
        temp = sheet.cell(row=i, column=1).value
        dates.append(temp.date())
        totals.append(sheet.cell(row=i, column=2).value)

    x = mdates.date2num(dates)
    y = totals

    poly_degree = 3
    z4 = np.polyfit(x, totals, poly_degree)
    p4 = np.poly1d(z4)

    print(p4)

    fig, cx = plt.subplots()

    xx = np.linspace(x.min(), x.max()+extrapolation, 1000)
    dd = mdates.num2date(xx)

    cx.plot(dd, p4(xx), '-g')
    cx.plot(dates, y, '+', color='b', label='blub')

    cx.grid()
    cx.set_ylim(0, 1000000)
    plt.show()


def example_function(*args):
    global filename
    print(args)
    filename = str(args)


def button_func(*args):
    print("pressed")


class HomeScreen(GridLayout):
    def __init__(self, **kwargs):
        super(HomeScreen, self).__init__(**kwargs)
        self.cols = 3
        self.rows = 3

        self.file_chooser = FileChooserIconView(path = '/home/')

        self.go_button = Button(text="Extrapolate!", on_press=plotValues)
        self.add_widget(self.go_button)

        self.add_widget(Label(text='File location'))
        self.username = TextInput(multiline=False)
        #self.username.bind(text=example_function)
        self.add_widget(self.username)


class MyApp(App):
    def build(self):
        return HomeScreen()


if __name__ == "__main__":
    #plotValues()
    MyApp().run()
