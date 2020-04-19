import openpyxl
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime
from datetime import timedelta
from datetime import date
import sys

error_string = {
    "bad_file_name":    "Invalid file name",
    "bad_file_format":  "File is incorrectly formatted",
    "bad_interp":       "Interpolation processing error",
    "pyplot_failure":   "Plotting failed to run properly",

    "unknown_error":    "Unknown error: "
}

def __get_future_date(base_date, additional_days):
    future_date = base_date + timedelta(additional_days)
    return future_date

def plot_values(file_path, degree=2, extrapolated_days=0):
    dates = []
    totals = []

    try:
        wb = openpyxl.load_workbook(file_path)
    except openpyxl.utils.exceptions.InvalidFileException:
        raise Exception(error_string["bad_file_name"])

    sheet = wb.get_sheet_by_name('Sheet1')

    try:
        for i in range(2, sheet.max_row):
            temp = sheet.cell(row=i, column=1).value
            dates.append(temp.date())
            totals.append(sheet.cell(row=i, column=2).value)

    except TypeError:
        raise Exception(error_string["bad_file_format"])

    except:
        e = sys.exc_info()[0]
        raise Exception(error_string["unknown_error"] + str(e))

    try:
        x_data = mdates.date2num(dates)
        y_data = totals

        # Generate polynomial
        polynomial_coefficients = np.polyfit(x_data, y_data, degree)
        f = np.poly1d(polynomial_coefficients)

        fig, cx = plt.subplots()

        if extrapolated_days > 0:
            future_date = mdates.date2num(__get_future_date(dates[-1], extrapolated_days))
            difference = future_date - x_data.max()

        x = np.linspace(x_data.min(), x_data.max() + difference)
        datetime_dates = mdates.num2date(x)

    except:
        e = sys.exc_info()[0]
        print("Error: " + str(e))
        raise Exception(error_string["bad_interp"])

    try:
        cx.plot(dates, y_data, '.k')
        cx.plot(datetime_dates, f(x), '-g')
        cx.grid()
        cx.set_ylim(0, f(x).max())
        plt.show()

    except:
        e = sys.exc_info()[0]
        print("Error: " + str(e))
        raise Exception(error_string["pyplot_failure"])

